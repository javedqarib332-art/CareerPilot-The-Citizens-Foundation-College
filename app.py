"""
TCF Discovery Agent — Flask App
Stage 1 of the AI Career Guidance System

Developed by Qarib Javed
"""

import os
from functools import wraps
from datetime import datetime
from io import BytesIO
from flask import Flask, render_template, request, jsonify, Response, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill

from discovery_engine import (
    get_all_questions,
    parse_submission,
    run_discovery_assessment,
)
import database
import field_data

app = Flask(__name__)
database.init_db()

# Change this before the pilot — this is what protects the counsellor dashboard.
# Better practice: set it as an environment variable instead of hardcoding.
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "tcf-counsellor-2026")


def require_dashboard_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.authorization
        if not auth or auth.password != DASHBOARD_PASSWORD:
            return Response(
                "Authentication required.", 401,
                {"WWW-Authenticate": 'Basic realm="Counsellor Dashboard"'}
            )
        return f(*args, **kwargs)
    return wrapper


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/questions")
def api_questions():
    return jsonify(get_all_questions())


@app.route("/api/submit", methods=["POST"])
def api_submit():
    payload = request.get_json(force=True)
    try:
        student = parse_submission(payload)
        result = run_discovery_assessment(student)

        skills_ratings_serializable = {
            k: {"rating": v[0], "reason": v[1]} for k, v in student.skills_ratings.items()
        }
        submission_id = database.save_submission(
            student_name=student.student_name,
            skills_ratings=skills_ratings_serializable,
            result=result,
            academic_ratings=student.academic_ratings,
            roll_number=student.roll_number,
            student_class=student.student_class,
        )
        result["submission_id"] = submission_id
        result["skills_ratings"] = [
            [k, v[0], v[1]] for k, v in student.skills_ratings.items()
        ]
        result["academic_ratings"] = student.academic_ratings

        result["suggested_fields_detail"] = [
            {
                "name": name,
                "slug": field_data.get_slug_for_field_name(name),
                "available": field_data.has_profile(name),
            }
            for name in result.get("suggested_fields", [])
        ]

        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/dashboard")
@require_dashboard_auth
def dashboard():
    submissions = database.get_all_submissions()
    total = len(submissions)
    valid_count = sum(1 for s in submissions if s["valid_response"])
    flagged_count = total - valid_count
    stats = {"total": total, "valid": valid_count, "flagged": flagged_count}
    return render_template("dashboard.html", submissions=submissions, stats=stats)


@app.route("/dashboard/export")
@require_dashboard_auth
def dashboard_export():
    submissions = database.get_all_submissions_full()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discovery Submissions"

    headers = [
        "ID", "Student Name", "Roll Number", "Class", "Submitted At",
        "Realistic", "Investigative", "Artistic", "Social", "Enterprising", "Conventional",
        "Suggested Fields", "Valid Response", "Contradictions Flagged",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="16233F", end_color="16233F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for s in submissions:
        r = s["riasec_scores"]
        ws.append([
            s["id"],
            s["student_name"],
            s["roll_number"],
            s["student_class"],
            s["created_at"][:16].replace("T", " "),
            r.get("R", ""), r.get("I", ""), r.get("A", ""), r.get("S", ""), r.get("E", ""), r.get("C", ""),
            ", ".join(s["suggested_fields"]),
            "Yes" if s["valid_response"] else "No",
            len(s["contradiction_flags"]),
        ])

    # Auto-fit column widths roughly based on content length
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"TCF_Discovery_Submissions_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/dashboard/<int:submission_id>")
@require_dashboard_auth
def dashboard_detail(submission_id):
    submission = database.get_submission_by_id(submission_id)
    if submission is None:
        return "Submission not found.", 404
    return render_template("dashboard_detail.html", s=submission)


@app.route("/fields")
def fields_directory():
    grouped = field_data.get_directory_data()
    total_count = sum(len(v) for v in grouped.values())
    return render_template("fields_directory.html", grouped=grouped, total_count=total_count)


@app.route("/field/<slug>")
def field_detail(slug):
    profile = field_data.get_profile_by_slug(slug)
    if profile is None:
        return render_template("field_not_ready.html"), 404
    return render_template("field_detail.html", f=profile)


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
